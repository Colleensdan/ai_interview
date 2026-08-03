"""Read the codebook (one row per code, plus its definition)."""

from __future__ import annotations

import io
from dataclasses import dataclass
from pathlib import Path

import openpyxl

from pipeline.names import clean_code_name

# The definition column is headed differently by different exports: the template
# codebook uses "Definition", the ATLAS.ti one uses "Comment".
DEFINITION_HEADERS = ("definition", "comment", "description")


class CodebookError(RuntimeError):
    """The codebook could not be read as a codebook."""


@dataclass(frozen=True)
class Code:
    name: str
    definition: str
    dimension: str  # source sheet name (a grouping in multi-sheet codebooks)


def _norm(s) -> str:
    return str(s).strip() if s is not None else ""


def load_codebook(path: str | Path) -> list[Code]:
    """Return every (code, definition) pair, in sheet then row order.

    Expects a header row with a "Code" column and a definition column (any of
    :data:`DEFINITION_HEADERS`). Robust to extra/blank rows and to the columns
    being in any order.
    """
    return _load_codebook_wb(openpyxl.load_workbook(path, data_only=True, read_only=True))


def load_codebook_bytes(data: bytes) -> list[Code]:
    """Same as :func:`load_codebook` but from in-memory bytes (RAM-only mode)."""
    return _load_codebook_wb(
        openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True))


def _definition_column(header: list[str]) -> int | None:
    for candidate in DEFINITION_HEADERS:
        if candidate in header:
            return header.index(candidate)
    return None


def _load_codebook_wb(wb) -> list[Code]:
    codes: list[Code] = []
    skipped: list[str] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            skipped.append(f"{sheet} (empty)")
            continue
        header = [_norm(c).lower() for c in rows[0]]
        def_col = _definition_column(header)
        if "code" not in header or def_col is None:
            skipped.append(f"{sheet} (headers: {', '.join(h for h in header if h) or 'none'})")
            continue
        code_col = header.index("code")
        for row in rows[1:]:
            if row is None:
                continue
            name = clean_code_name(row[code_col]) if code_col < len(row) else ""
            definition = _norm(row[def_col]) if def_col < len(row) else ""
            if name:
                codes.append(
                    Code(name=name, definition=definition, dimension=sheet)
                )
    wb.close()
    if not codes:
        # Previously this returned [] and the whole pipeline ran to a clean exit
        # having coded nothing at all. Refuse instead.
        raise CodebookError(
            "No codes found. Every sheet was skipped for want of a 'Code' column "
            f"plus one of {DEFINITION_HEADERS}. Sheets seen: {'; '.join(skipped) or 'none'}"
        )
    return codes

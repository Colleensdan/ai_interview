"""Read the human ground-truth code x document count matrix (ATLAS.ti export).

This is the CodeDocumentTable: row 1 holds document headers like
``27502M\\nGr=53`` or ``R_2Eu2LaCpcaFrPEP\\nGr=11``; each subsequent row is
``○ <code>\\nGr=NN`` followed by the integer count of that code in each
document. We return counts keyed by the canonical document key so they line up
with the interview filenames.

Two structural details matter. ATLAS.ti appends a ``Totals`` row and a
``Totals`` column, which are dropped — left in, they enter the analysis as a
phantom code and a phantom document. And the ``Gr=`` annotations are kept: they
are the export's own quotation counts, used to cross-check the quotes workbook
and to disambiguate quote sheets whose names collide after truncation.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from pipeline.names import clean_code_name, groundedness

TOTALS_LABEL = "totals"


@dataclass(frozen=True)
class GroundTruthTable:
    """The parsed count matrix plus the metadata other readers need."""

    counts: dict[str, dict[str, int]]
    doc_keys: list[str]
    code_groundedness: dict[str, int] = field(default_factory=dict)
    doc_groundedness: dict[str, int] = field(default_factory=dict)

    def document_for_number(self, n: int) -> str | None:
        """Map an ATLAS.ti document number to a document key.

        Quotation IDs are ``<document number>:<quotation number>``, and the
        column order of the count matrix is that same document numbering
        (1-based). This is what lets a quote be attributed to one specific
        transcript instead of being text-searched across all of them.
        """
        if 1 <= n <= len(self.doc_keys):
            return self.doc_keys[n - 1]
        return None


def _doc_header_key(label) -> str:
    """`'27502M\\nGr=53'` -> `'27502'`; `'R_2Eu…\\nGr=11'` -> `'R_2Eu…'`.

    Leading digits when the name has them (the template's ATLAS names truncate
    the filename), else the whole first line (the chat export's response IDs).
    """
    if label is None:
        return ""
    first = str(label).split("\n", 1)[0].strip()
    m = re.match(r"(\d+)", first)
    return m.group(1) if m else first


def load_ground_truth_counts(
    path: str | Path, sheet: str
) -> tuple[dict[str, dict[str, int]], list[str]]:
    """Return (counts, doc_keys). See :func:`load_ground_truth_table`."""
    table = load_ground_truth_table(path, sheet)
    return table.counts, table.doc_keys


def load_ground_truth_counts_bytes(
    data: bytes, sheet: str
) -> tuple[dict[str, dict[str, int]], list[str]]:
    """Same as :func:`load_ground_truth_counts` but from in-memory bytes."""
    table = load_ground_truth_table_bytes(data, sheet)
    return table.counts, table.doc_keys


def load_ground_truth_table(path: str | Path, sheet: str) -> GroundTruthTable:
    """Parse the full count matrix, Totals dropped and Gr= annotations kept."""
    return _table_from_wb(
        openpyxl.load_workbook(path, data_only=True, read_only=True), sheet)


def load_ground_truth_table_bytes(data: bytes, sheet: str) -> GroundTruthTable:
    """Same as :func:`load_ground_truth_table` but from in-memory bytes."""
    return _table_from_wb(
        openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True), sheet)


def _table_from_wb(wb, sheet: str) -> GroundTruthTable:
    if sheet not in wb.sheetnames:
        wb.close()
        raise KeyError(
            f"Sheet {sheet!r} not in the count matrix (found: {', '.join(wb.sheetnames)}). "
            "Set AICODE_GT_SHEET if the export uses a different name."
        )
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return GroundTruthTable({}, [])

    # Column 0 is the code label; columns 1.. are documents, less the Totals column.
    header = rows[0]
    doc_cols = [
        (i, _doc_header_key(h), groundedness(h))
        for i, h in enumerate(header[1:], start=1)
        if _doc_header_key(h) and _doc_header_key(h).lower() != TOTALS_LABEL
    ]
    doc_keys = [key for _, key, _ in doc_cols]
    doc_gr = {key: gr for _, key, gr in doc_cols if gr is not None}

    counts: dict[str, dict[str, int]] = {}
    code_gr: dict[str, int] = {}
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        code = clean_code_name(row[0])
        if not code or code.lower() == TOTALS_LABEL:
            continue
        per_doc: dict[str, int] = {}
        for i, key, _ in doc_cols:
            value = row[i] if i < len(row) else None
            try:
                per_doc[key] = int(value) if value is not None else 0
            except (TypeError, ValueError):
                per_doc[key] = 0
        counts[code] = per_doc
        gr = groundedness(row[0])
        if gr is not None:
            code_gr[code] = gr
    return GroundTruthTable(counts, doc_keys, code_gr, doc_gr)

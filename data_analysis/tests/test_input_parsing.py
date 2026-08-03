"""Tests for the input readers, against the shapes the real exports actually use.

Each case here corresponds to a way the pipeline previously failed *silently* on
the study data: a codebook whose definition column is headed "Comment" loaded
zero codes; `.txt` transcripts in a nested folder loaded zero interviews; a
`○` bullet survived into the code name; and the `Totals` row and column entered
the analysis as a phantom code and a phantom document.
"""

from __future__ import annotations

import openpyxl
import pytest

from pipeline.codebook import CodebookError, load_codebook
from pipeline.ground_truth import load_ground_truth_table
from pipeline.interviews import (
    NoTranscriptsError,
    annotate_roles,
    has_explicit_roles,
    load_interviews,
    parse_turns,
    select_sample,
)
from pipeline.names import clean_code_name, groundedness
from pipeline.quote_sheets import build_quote_index


def _codebook(tmp_path, header, rows, name="Codebook.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Codes"
    ws.append(header)
    for r in rows:
        ws.append(r)
    path = tmp_path / name
    wb.save(path)
    return path


def _counts(tmp_path, header, rows, name="Counts.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CodeDocumentTable"
    ws.append(header)
    for r in rows:
        ws.append(r)
    path = tmp_path / name
    wb.save(path)
    return path


# --- code-name normalisation -------------------------------------------------

def test_clean_code_name_handles_every_export_spelling():
    assert clean_code_name("○ anger/irritation\nGr=17") == "anger/irritation"
    assert clean_code_name("● A: alienating\nGr=26") == "A: alienating"
    assert clean_code_name("anger\\\\/irritation") == "anger/irritation"
    assert clean_code_name(None) == ""


def test_groundedness_is_read_off_the_label():
    assert groundedness("○ worry\nGr=42") == 42
    assert groundedness("worry") is None


# --- codebook ----------------------------------------------------------------

def test_codebook_accepts_comment_as_the_definition_column(tmp_path):
    path = _codebook(tmp_path, ["Code", "Comment"], [["worry", "feeling worried"]])
    codes = load_codebook(path)
    assert [(c.name, c.definition) for c in codes] == [("worry", "feeling worried")]


def test_codebook_still_accepts_definition(tmp_path):
    path = _codebook(tmp_path, ["Code", "Freq", "Definition"], [["worry", 3, "d"]])
    assert load_codebook(path)[0].definition == "d"


def test_codebook_unescapes_slashes(tmp_path):
    path = _codebook(tmp_path, ["Code", "Comment"], [["anger\\\\/irritation", "d"]])
    assert load_codebook(path)[0].name == "anger/irritation"


def test_unrecognised_codebook_raises_instead_of_loading_nothing(tmp_path):
    """The failure that produced a clean run which coded nothing at all."""
    path = _codebook(tmp_path, ["Label", "Notes"], [["worry", "d"]])
    with pytest.raises(CodebookError):
        load_codebook(path)


# --- count matrix ------------------------------------------------------------

def test_totals_row_and_column_are_dropped(tmp_path):
    path = _counts(
        tmp_path,
        [None, "R_aaa\nGr=2", "R_bbb\nGr=1", "Totals"],
        [["○ worry\nGr=3", 2, 1, 3], ["Totals", 2, 1, 3]],
    )
    table = load_ground_truth_table(path, "CodeDocumentTable")
    assert list(table.counts) == ["worry"]
    assert table.doc_keys == ["R_aaa", "R_bbb"]
    assert "Totals" not in table.counts["worry"]


def test_groundedness_and_document_numbering_are_captured(tmp_path):
    path = _counts(
        tmp_path,
        [None, "R_aaa\nGr=2", "R_bbb\nGr=1"],
        [["○ worry\nGr=3", 2, 1]],
    )
    table = load_ground_truth_table(path, "CodeDocumentTable")
    assert table.code_groundedness["worry"] == 3
    assert table.document_for_number(1) == "R_aaa"
    assert table.document_for_number(2) == "R_bbb"
    assert table.document_for_number(3) is None


def test_missing_sheet_names_the_sheets_it_did_find(tmp_path):
    path = _counts(tmp_path, [None, "R_aaa"], [["worry", 1]])
    with pytest.raises(KeyError, match="CodeDocumentTable"):
        load_ground_truth_table(path, "NotThere")


# --- transcripts -------------------------------------------------------------

def test_txt_transcripts_load_from_a_nested_directory(tmp_path):
    nested = tmp_path / "All chats" / "All chats"
    nested.mkdir(parents=True)
    (nested / "R_aaa.txt").write_text("user: hello\n", encoding="utf-8")
    (nested / "R_aaa.txt:Zone.Identifier").write_text("junk", encoding="utf-8")
    interviews = load_interviews(tmp_path)
    assert [iv.title for iv in interviews] == ["R_aaa.txt"]
    assert interviews[0].key == "R_aaa"


def test_empty_transcript_directory_raises(tmp_path):
    with pytest.raises(NoTranscriptsError):
        load_interviews(tmp_path)


def test_a_line_without_a_role_prefix_continues_the_previous_turn():
    """81 of the 104 real chats contain wrapped multi-line messages."""
    text = ("assistant: First question.\n"
            "Still the assistant, wrapped onto a new line.\n"
            "user: My answer.\n"
            "Still me.")
    turns = parse_turns(text)
    assert len(turns) == 2
    assert "wrapped onto a new line" in turns[0][2]
    assert "Still me" in turns[1][2]


def test_roles_are_classified_and_tagged():
    text = "assistant: A question?\nuser: An answer."
    assert has_explicit_roles(text)
    tagged = annotate_roles(text)
    assert "[INTERVIEWER — do NOT code this]" in tagged
    assert "[INTERVIEWEE — you MAY code this]" in tagged
    # Content must survive verbatim or quote highlighting breaks.
    assert "An answer." in tagged


def test_positional_speakers_are_left_untagged():
    text = "Speaker 1: A question?\nSpeaker 2: An answer."
    assert not has_explicit_roles(text)
    assert annotate_roles(text) == text


def test_sample_fraction_is_read_at_call_time(monkeypatch, tmp_path):
    import config
    nested = tmp_path / "chats"
    nested.mkdir()
    for i in range(10):
        (nested / f"R_{i:02d}.txt").write_text("user: hi\n", encoding="utf-8")
    interviews = load_interviews(nested)
    monkeypatch.setattr(config, "SAMPLE_FRACTION", 0.5)
    assert len(select_sample(interviews)) == 5
    monkeypatch.setattr(config, "SAMPLE_FRACTION", 1.0)
    assert len(select_sample(interviews)) == 10


# --- quote sheets ------------------------------------------------------------

def _quotes_wb(sheets):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name[:31])
        ws.append(["ID", "Quotation Content"])
        for r in rows:
            ws.append(r)
    return wb


def test_truncated_colliding_sheets_are_told_apart_by_groundedness():
    """The real collision: three codes, three sheets, names give no clue.

    Sheet order does not follow code order, so anything positional or fuzzy
    assigns these wrongly and silently merges three distinct codes' quotes.
    """
    codes = [
        "pro-environmental social norms increased",
        "pro-environmental social norms reduced",
        "pro-environmental social norms unchanged/continues",
    ]
    wb = _quotes_wb({
        "pro-environmental social norm1": [["1:1", "a"], ["1:2", "b"]],
        "pro-environmental social norm2": [["2:1", "c"]],
        "pro-environmental social norms": [["3:1", "d"], ["3:2", "e"], ["3:3", "f"]],
    })
    index = build_quote_index(
        wb, codes,
        groundedness={codes[0]: 3, codes[1]: 2, codes[2]: 1},
    )
    assert index.sheet_to_code["pro-environmental social norm1"] == codes[1]
    assert index.sheet_to_code["pro-environmental social norm2"] == codes[2]
    assert index.sheet_to_code["pro-environmental social norms"] == codes[0]
    assert not index.ambiguous and not index.unresolved_sheets


def test_quotes_are_attributed_to_their_own_document():
    wb = _quotes_wb({"worry": [["2:5", "a shared phrase"]]})
    index = build_quote_index(
        wb, ["worry"], doc_for_number={1: "R_aaa", 2: "R_bbb"}.get)
    assert index.by_code["worry"][0].doc == "R_bbb"


def test_metadata_sheets_are_not_treated_as_codes():
    wb = _quotes_wb({"worry": [["1:1", "q"]], "Info": [["x", "y"]]})
    index = build_quote_index(wb, ["worry"], non_code_sheets={"info"})
    assert list(index.by_code) == ["worry"]
    assert not index.unresolved_sheets


def test_two_sheets_claiming_one_code_is_reported_not_merged():
    wb = _quotes_wb({"worry": [["1:1", "a"]], "worry ": [["2:1", "b"]]})
    index = build_quote_index(wb, ["worry"])
    assert index.ambiguous
    assert len(index.by_code.get("worry", [])) == 1


def test_code_with_counts_but_no_sheet_is_reported():
    """The dual-source gap: human-coded, but nothing to highlight."""
    wb = _quotes_wb({"worry": [["1:1", "q"]]})
    index = build_quote_index(
        wb, ["worry", "summary"], groundedness={"worry": 1, "summary": 57})
    assert index.codes_without_sheets == ["summary"]

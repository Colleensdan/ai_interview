"""Read Task 1 outputs and shape them for the Task 2 UI.

Sources:
- coding.db: AI coding results + per-code Cohen's kappa (per model).
- transcripts (.docx): the interview text shown in both panels.
- Ground Truth.xlsx: human-coded quotes, located into transcripts for highlighting.
- codebook + DefinitionStore: code definitions (current, possibly edited).
"""

from __future__ import annotations

import re
import sqlite3
from difflib import get_close_matches
from functools import lru_cache

import openpyxl

import config
from pipeline.codebook import load_codebook
from pipeline.interviews import load_interviews
from .quotes import build_norm, locate, merge_overlaps

MAJORITY = "majority_vote"


def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", s.lower())


def code_color(code_name: str) -> str:
    """Deterministic, well-spaced HSL colour per code (stable across panels)."""
    h = (abs(hash(_norm(code_name))) % 360)
    return f"hsl({h}, 70%, 78%)"


class DataStore:
    def __init__(self) -> None:
        self.db_path = str(config.DB_PATH)
        self._interviews = {iv.title: iv.text for iv in load_interviews(config.INTERVIEWS_DIR)}
        self._codes = load_codebook(config.CODEBOOK_PATH)  # base defs (overridden by store)
        self._gt_quotes = self._load_ground_truth_quotes()
        # Cache normalized transcript per doc (built once, reused by locate()).
        self._norm: dict[str, tuple[str, list[int]]] = {
            doc: build_norm(text) for doc, text in self._interviews.items()
        }
        # Cache located human hits per doc (all codes); GT quotes never change.
        self._human_cache: dict[str, list[dict]] = {}

    # --- connections -------------------------------------------------------
    def _conn(self) -> sqlite3.Connection:
        import db as _db

        return _db.connect(self.db_path, row_factory=True)

    # --- codebook ----------------------------------------------------------
    def codebook(self) -> list:
        return self._codes

    def base_definitions(self) -> dict[str, str]:
        return {c.name: c.definition for c in self._codes}

    # --- models ------------------------------------------------------------
    def live_model(self) -> str | None:
        """The model the live Azure deployment would code under (or None)."""
        if config.AZURE_DEPLOYMENT:
            return f"azure_openai-{config.AZURE_DEPLOYMENT}"
        return None

    def models(self) -> list[str]:
        """Selectable 'model data' for the UI, most-recently-used first.

        Ordered by latest run (not alphabetically) so the active model leads;
        majority vote is always last.
        """
        with self._conn() as c:
            rows = c.execute(
                "SELECT model, MAX(id) AS last FROM kappa GROUP BY model ORDER BY last DESC"
            ).fetchall()
        ordered = [r["model"] for r in rows]
        reals = [m for m in ordered if m != MAJORITY]
        return reals + ([MAJORITY] if MAJORITY in ordered else [])

    def default_model(self) -> str:
        """Prefer the live deployment's model if it's in the data, else the
        most-recently-used real model (deterministic, not alphabetical)."""
        ms = self.models()
        live = self.live_model()
        if live and live in ms:
            return live
        return ms[0] if ms else MAJORITY

    # --- kappa -------------------------------------------------------------
    def latest_kappa(self, model: str) -> dict[str, float | None]:
        """Most recent kappa per code for a model (reflects re-analyses)."""
        with self._conn() as c:
            rows = c.execute(
                "SELECT code_name, kappa FROM kappa k WHERE model=? AND id=("
                "  SELECT MAX(id) FROM kappa WHERE model=? AND code_name=k.code_name)",
                (model, model),
            ).fetchall()
        return {r["code_name"]: r["kappa"] for r in rows}

    def overview(self, model: str) -> dict:
        kappas = self.latest_kappa(model)
        target = config.KAPPA_TARGET
        success, fail = [], []
        for code in (c.name for c in self._codes):
            k = kappas.get(code)
            entry = {"code": code, "kappa": k}
            if k is not None and k > target:
                success.append(entry)
            else:
                fail.append(entry)
        total = len(success) + len(fail)
        success.sort(key=lambda e: e["code"])
        fail.sort(key=lambda e: (e["kappa"] is None, e["kappa"] if e["kappa"] is not None else 0))
        return {
            "model": model,
            "target": target,
            "total": total,
            "n_success": len(success),
            "n_fail": len(fail),
            "pct_success": round(100 * len(success) / total, 1) if total else 0.0,
            "pct_fail": round(100 * len(fail) / total, 1) if total else 0.0,
            "success": success,
            "fail": fail,
        }

    # --- documents ---------------------------------------------------------
    def sampled_docs(self) -> list[str]:
        """Docs coded by AI (i.e. labelled by AI AND human — the review set)."""
        with self._conn() as c:
            rows = c.execute(
                "SELECT DISTINCT document_title FROM coding_results ORDER BY document_title"
            ).fetchall()
        docs = [r["document_title"] for r in rows]
        return [d for d in docs if d in self._interviews]

    # --- AI coding ---------------------------------------------------------
    def ai_hits(self, model: str, doc: str, codes: list[str] | None = None) -> list[dict]:
        """Latest AI (code, quote, reason) rows for a doc.

        For majority vote there are no per-quote rows, so fall back to the
        first real model's quotes (the vote is about presence, not text).
        """
        src_model = model if model != MAJORITY else self.default_model()
        params = [src_model, doc]
        clause = ""
        if codes:
            clause = f" AND code_name IN ({','.join('?' * len(codes))})"
            params += codes
        with self._conn() as c:
            rows = c.execute(
                "SELECT code_name, quote, reason FROM coding_results r "
                "WHERE model=? AND document_title=?" + clause +
                " AND run_id=(SELECT run_id FROM coding_results "
                "            WHERE model=r.model AND code_name=r.code_name "
                "            ORDER BY id DESC LIMIT 1)",
                params,
            ).fetchall()
        return [dict(r) for r in rows]

    # --- ground truth ------------------------------------------------------
    def _load_ground_truth_quotes(self) -> dict[str, list[str]]:
        """Map code name -> list of human-coded quote strings (from Ground Truth.xlsx)."""
        wb = openpyxl.load_workbook(config.GROUND_TRUTH_QUOTES_PATH, data_only=True, read_only=True)
        code_names = [c.name for c in self._codes]
        norm_to_code = {_norm(n): n for n in code_names}
        result: dict[str, list[str]] = {}
        for sheet in wb.sheetnames:
            code = norm_to_code.get(_norm(sheet))
            if code is None:
                close = get_close_matches(_norm(sheet), list(norm_to_code), n=1, cutoff=0.8)
                code = norm_to_code[close[0]] if close else sheet
            ws = wb[sheet]
            quotes = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) > 1 and row[1]:
                    quotes.append(str(row[1]))
            result.setdefault(code, []).extend(quotes)
        wb.close()
        return result

    def _all_human_hits(self, doc: str) -> list[dict]:
        """All human-coded quotes located in this doc (cached; computed once)."""
        if doc in self._human_cache:
            return self._human_cache[doc]
        text = self._interviews.get(doc, "")
        norm = self._norm.get(doc)
        hits = []
        for code, quotes in self._gt_quotes.items():
            for q in quotes:
                span = locate(text, q, norm=norm)
                if span:
                    hits.append({"code": code, "quote": q, "start": span.start, "end": span.end})
        self._human_cache[doc] = hits
        return hits

    def human_hits(self, doc: str, codes: list[str] | None = None) -> list[dict]:
        """Human-coded quotes located in this doc, optionally filtered by code."""
        hits = self._all_human_hits(doc)
        if codes is None:
            return hits
        want = set(codes)
        return [h for h in hits if h["code"] in want]

    # --- interview view ----------------------------------------------------
    def interview_view(self, doc: str, model: str, codes: list[str] | None = None,
                       definitions: dict[str, str] | None = None) -> dict:
        text = self._interviews.get(doc, "")
        norm = self._norm.get(doc)
        defs = definitions or self.base_definitions()

        # AI highlights: locate each AI quote in the text.
        ai_spans = []
        for hit in self.ai_hits(model, doc, codes):
            span = locate(text, hit["quote"], norm=norm)
            if span:
                ai_spans.append((span.start, span.end,
                                 {"code": hit["code_name"], "reason": hit["reason"]}))
        human_spans = [(h["start"], h["end"], {"code": h["code"]})
                       for h in self.human_hits(doc, codes)]

        used_codes = sorted({p["code"] for _, _, p in ai_spans + human_spans})
        return {
            "title": doc,
            "ai_segments": self._segments(text, ai_spans),
            "human_segments": self._segments(text, human_spans),
            "colors": {c: code_color(c) for c in used_codes},
            "meanings": {c: defs.get(c, "") for c in used_codes},
            "ai_codes": sorted({p["code"] for _, _, p in ai_spans}),
            "human_codes": sorted({p["code"] for _, _, p in human_spans}),
        }

    @staticmethod
    def _segments(text: str, spans: list[tuple[int, int, dict]]) -> list[dict]:
        """Turn char spans into ordered text segments for rendering."""
        merged = merge_overlaps(spans)
        segments = []
        cursor = 0
        for start, end, payloads in merged:
            if start > cursor:
                segments.append({"text": text[cursor:start], "codes": []})
            codes = []
            for pl in payloads:
                entry = {"code": pl["code"]}
                if "reason" in pl:
                    entry["reason"] = pl["reason"]
                codes.append(entry)
            segments.append({"text": text[start:end], "codes": codes})
            cursor = end
        if cursor < len(text):
            segments.append({"text": text[cursor:], "codes": []})
        return segments

    # --- failure modes (spec 5.3) -----------------------------------------
    def failures(self, model: str, codes: list[str]) -> list[dict]:
        """Per code: false positives (AI present, human absent) and false
        negatives (human present, AI absent), per document, with reasons."""
        docs = self.sampled_docs()
        out = []
        for code in codes:
            fp, fn = [], []
            for doc in docs:
                ai = self.ai_hits(model, doc, [code])
                human = self.human_hits(doc, [code])
                ai_present, human_present = bool(ai), bool(human)
                if ai_present and not human_present:
                    fp.append({"document": doc,
                               "ai_quotes": [{"quote": h["quote"], "reason": h["reason"]} for h in ai]})
                elif human_present and not ai_present:
                    fn.append({"document": doc,
                               "human_quotes": [h["quote"] for h in human]})
            out.append({"code": code, "false_positives": fp, "false_negatives": fn})
        return out
